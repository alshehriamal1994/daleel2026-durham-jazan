# Score the filled-in blind check sheet against the hidden key.
# Usage: python3 score_human_eval.py synthetic_data_blind_check_FILLED.xlsx
import sys, json
import numpy as np
from openpyxl import load_workbook

LABELS = ["AS", "AN", "ST", "TE", "CO", "OT"]
key = json.load(open(sys.argv[2] if len(sys.argv) > 2 else "human_eval_key.json", encoding="utf-8"))
wb = load_workbook(sys.argv[1] if len(sys.argv) > 1 else "synthetic_data_blind_check.xlsx", data_only=True)

def kappa(g, a):
    po = (g == a).mean(); pe = g.mean() * a.mean() + (1 - g.mean()) * (1 - a.mean())
    return (po - pe) / (1 - pe) if pe < 1 else 1.0

# ---- Task 1 (paragraph-level multi-label), same metrics as paper_analysis/synth_agreement.py
ws = wb["Task1"] if "Task1" in wb.sheetnames else wb["Paragraphs"]; human, nat1 = {}, []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None: continue
    no = int(row[0]); labs = {LABELS[k] for k in range(6) if str(row[3 + k] or "").strip().lower() == "x"}
    human[no] = labs
    if row[9] not in (None, ""): nat1.append(float(row[9]))
gen = {k["no"]: set(k["labels"]) for k in key["task1"]}
done = [n for n in gen if n in human and (human[n] or True)]
filled = [n for n in gen if human.get(n)]
print(f"Task 1: {len(filled)}/{len(gen)} paragraphs labelled by the human")
ids = filled
if ids:
    exact = np.mean([gen[i] == human[i] for i in ids])
    tp = fp = fn = 0
    print(f"  exact label-set agreement: {exact:.3f}")
    print(f"  {'label':<6}{'n_gen':>6}{'agree%':>8}{'kappa':>7}{'P':>7}{'R':>7}{'F1':>7}")
    for lab in LABELS:
        g = np.array([lab in gen[i] for i in ids]); a = np.array([lab in human[i] for i in ids])
        t = int((g & a).sum()); f_ = int((~g & a).sum()); n_ = int((g & ~a).sum()); tp += t; fp += f_; fn += n_
        P = t / (t + f_) if t + f_ else 0; R = t / (t + n_) if t + n_ else 0; F = 2 * P * R / (P + R) if P + R else 0
        print(f"  {lab:<6}{int(g.sum()):>6}{(g==a).mean():>8.3f}{kappa(g,a):>7.3f}{P:>7.3f}{R:>7.3f}{F:>7.3f}")
    P = tp / (tp + fp) if tp + fp else 0; R = tp / (tp + fn) if tp + fn else 0
    print(f"  micro-F1 (human vs generator): {2*P*R/(P+R) if P+R else 0:.3f}   (P = generator labels confirmed by human: {P:.3f}; R = human labels present in generator: {R:.3f})")
if nat1: print(f"  naturalness: mean {np.mean(nat1):.2f}, share rated 3 = {np.mean([x==3 for x in nat1]):.2f}, rated 1 = {np.mean([x==1 for x in nat1]):.2f}")

# ---- Task 2 (segment-level single label)
ws = wb["Task2"] if "Task2" in wb.sheetnames else None; hl, nat2 = {}, []
if ws is None: sys.exit(0)
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None: continue
    no, seg = int(row[0]), int(row[1]); lab = str(row[5] or "").strip().upper()
    if lab: hl[(no, seg)] = lab
    if row[6] not in (None, ""): nat2.append(float(row[6]))
segs = [(k["no"], k["seg"], k["label"]) for k in key["task2"]]
sc = [(n, s, g) for n, s, g in segs if (n, s) in hl]
print(f"\nTask 2: {len(sc)}/{len(segs)} segments labelled by the human")
if sc:
    acc = np.mean([hl[(n, s)] == g for n, s, g in sc]); print(f"  segment label agreement (accuracy): {acc:.3f}")
    print(f"  {'label':<6}{'n_gen':>6}{'confirmed%':>12}   human said instead")
    for lab in LABELS:
        rows = [(n, s, g) for n, s, g in sc if g == lab]
        if not rows: continue
        conf = np.mean([hl[(n, s)] == lab for n, s, g in rows])
        from collections import Counter
        other = Counter(hl[(n, s)] for n, s, g in rows if hl[(n, s)] != lab)
        print(f"  {lab:<6}{len(rows):>6}{conf:>12.3f}   {dict(other) if other else ''}")
    # Cohen's kappa over the 7-way segment labels
    g = [g for n, s, g in sc]; a = [hl[(n, s)] for n, s, g in sc]; cats = sorted(set(g) | set(a))
    po = acc; pe = sum((np.mean([x == c for x in g])) * (np.mean([x == c for x in a])) for c in cats)
    print(f"  Cohen's kappa (7-way): {(po-pe)/(1-pe):.3f}")
if nat2: print(f"  naturalness: mean {np.mean(nat2):.2f}, share rated 3 = {np.mean([x==3 for x in nat2]):.2f}, rated 1 = {np.mean([x==1 for x in nat2]):.2f}")
