# Small blind human check: 30 Task 1 synthetic paragraphs (15 editorial, 15 debate), labels hidden.
import json, os, random, re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = os.environ.get("DALEEL_ROOT", os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))  # repository root
LABELS = ["AS", "AN", "ST", "TE", "CO", "OT"]
rng = random.Random(2026)
t1 = []
for f in ["data/synth_all.jsonl", "data/synth_v2/t1_batch_agent.jsonl"]:
    for l in open(f"{ROOT}/{f}", encoding="utf-8"):
        t1.append(json.loads(l))
assert len(t1) == 291

def pick(n, strata):
    pools = [list(s) for s in strata]
    for p in pools: rng.shuffle(p)
    out, k = [], 0
    while len(out) < n and any(pools):
        p = pools[k % len(pools)]; k += 1
        while p:
            i = p.pop()
            if i not in out: out.append(i); break
    return out
def strata(typ):
    idx = [i for i, r in enumerate(t1) if r["type"] == typ]
    return [[i for i in idx if "ST" in t1[i]["labels"]], [i for i in idx if "CO" in t1[i]["labels"]],
            [i for i in idx if "OT" in t1[i]["labels"]], idx]
sel = pick(15, strata("editorial")) + pick(15, strata("debate")); rng.shuffle(sel)

prompts = open(f"{ROOT}/PROMPTS.md", encoding="utf-8").read()
m = re.search(r"التعريفات:\s*(.*?)\nأخرج فقط", prompts, re.S); defs_ar = m.group(1).strip()
TYPE_AR = {"editorial": "مقال افتتاحي", "debate": "مناظرة"}

wb = Workbook(); ws = wb.active; ws.title = "Paragraphs"; ws.sheet_view.rightToLeft = True
ws.append(["No.", "Type", "Paragraph"] + LABELS + ["Natural? (1-3)"])
fill = PatternFill("solid", fgColor="DDEBF7")
for c in range(1, 11):
    cell = ws.cell(1, c); cell.font = Font(bold=True); cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.freeze_panes = "A2"
dv_x = DataValidation(type="list", formula1='"x"', allow_blank=True); dv_n = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
ws.add_data_validation(dv_x); ws.add_data_validation(dv_n)
key = []
for n, i in enumerate(sel, 1):
    r = t1[i]
    ws.append([n, TYPE_AR[r["type"]], r["text"]] + [""] * 6 + [""])
    row = ws.max_row
    ws.cell(row, 3).alignment = Alignment(wrap_text=True, horizontal="right", vertical="top", readingOrder=2)
    ws.cell(row, 3).font = Font(size=13)
    for c in (1, 2, 10): ws.cell(row, c).alignment = Alignment(horizontal="center", vertical="top")
    for c in range(4, 10):
        ws.cell(row, c).alignment = Alignment(horizontal="center", vertical="top"); dv_x.add(ws.cell(row, c))
    dv_n.add(ws.cell(row, 10))
    ws.row_dimensions[row].height = max(60, 18 * (len(r["text"]) // 65 + 1))
    key.append({"no": n, "source_id": i, "type": r["type"], "labels": r["labels"]})
widths = {1: 5, 2: 14, 3: 90, 10: 12}
for c in range(1, 11): ws.column_dimensions[get_column_letter(c)].width = widths.get(c, 6)

ws2 = wb.create_sheet("Definitions"); ws2.sheet_view.rightToLeft = True
ws2.append(["Label definitions (official guidelines)"]); ws2.cell(1, 1).font = Font(bold=True, size=12)
for d in defs_ar.split("\n"):
    ws2.append([d]); ws2.cell(ws2.max_row, 1).alignment = Alignment(wrap_text=True, horizontal="right", vertical="top", readingOrder=2)
    ws2.cell(ws2.max_row, 1).font = Font(size=13)
ws2.column_dimensions["A"].width = 120

out = "blind_check_30.xlsx"; wb.save(out)
json.dump({"seed": 2026, "task1": key, "task2": []}, open("human_eval_key_30.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)
from collections import Counter
print(f"{len(key)} paragraphs ({sum(k['type']=='editorial' for k in key)} ed / {sum(k['type']=='debate' for k in key)} db); label counts {dict(Counter(l for k in key for l in k['labels']))}")
